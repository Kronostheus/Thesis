import openpyxl as px
import pandas as pd
import glob

WORKBOOK_DIR = "Data/Workbooks/"
DATA_DIR = "Data/Coding_Schemes/"

COLS = ['C', 'F', 'I', 'L', 'O']
STYLE_VALUE = {'Bad': 0, 'Neutral': 0.5, 'Good': 1}

files = [file for file in glob.glob(WORKBOOK_DIR + '*.xlsx')]


def build_dataframe():
    """
    Build dataframe containing value associated with each choice
    :return: Dataframe
    """
    col_dict = {}

    for file in files:
        idx = str(files.index(file) + 1)
        sheet = px.load_workbook(file).active

        for col in COLS:
            col_dict[idx + col] = [STYLE_VALUE[sheet[col + str(cell_num)].style] for cell_num in range(2, sheet.max_row + 1)]

    return pd.DataFrame(col_dict)


sheet_df = build_dataframe()

# Get the 5 columns associated with each file
file_cols = [list(sheet_df.columns)[start:start+5] for start in range(0, sheet_df.shape[1] - 1, 5)]

percs = []


def print_out():
    """
    Not pretty but only here for printing
    :return:
    """
    for file_index in range(len(files)):
        cols = file_cols[file_index]

        good_first = sheet_df[cols][sheet_df[cols[0]] == 1].shape[0]

        good_sec = sheet_df[cols][(sheet_df[cols[0]] != 1) &
                                  (sheet_df[cols[1]] == 1)].shape[0]

        good_third = sheet_df[cols][(sheet_df[cols[0]] != 1) &
                                    (sheet_df[cols[1]] != 1) &
                                    (sheet_df[cols[2]] == 1)].shape[0]

        good_fourth = sheet_df[cols][(sheet_df[cols[0]] != 1) &
                                     (sheet_df[cols[1]] != 1) &
                                     (sheet_df[cols[2]] != 1) &
                                     (sheet_df[cols[3]] == 1)].shape[0]

        good_fifth = sheet_df[cols][(sheet_df[cols[0]] != 1) &
                                     (sheet_df[cols[1]] != 1) &
                                     (sheet_df[cols[2]] != 1) &
                                     (sheet_df[cols[3]] != 1) &
                                     (sheet_df[cols[4]] == 1)].shape[0]

        good_neither = sheet_df[cols][(sheet_df[cols[0]] != 1) &
                                     (sheet_df[cols[1]] != 1) &
                                     (sheet_df[cols[2]] != 1) &
                                     (sheet_df[cols[3]] != 1) &
                                     (sheet_df[cols[4]] != 1)].shape[0]

        perc = round(100 * (good_first + good_sec + good_third + good_fourth + good_fifth) / sheet_df.shape[0], 2)
        percs.append(perc)

        print("##################" + files[file_index] + "##################")

        print("A total of {} Good categories found in first choice\n"
              "Representing {}% of Good choices.\n".format(good_first,
                                                                   round(100 * (good_first)
                                                                         / sheet_df.shape[0], 2))
              )

        print("A total of {} Good categories found in second choice, while first either Bad or Neutral.\n"
              "If we were to replace these, the original {} Good choices would total {} out of {}.\n"
              "That would represent {}% of Good choices.\n".format(good_sec,
                                                                   good_first,
                                                                   good_first + good_sec,
                                                                   sheet_df.shape[0],
                                                                   round(100 * (good_first + good_sec)
                                                                         / sheet_df.shape[0], 2))
              )

        print("A total of {} Good categories found in third choice, while first either Bad or Neutral.\n"
              "If we were to replace these, the previous {} Good choices would total {} out of {}.\n"
              "That would represent {}% of Good choices.\n".format(good_third,
                                                                   good_first + good_sec,
                                                                   good_first + good_sec + good_third,
                                                                   sheet_df.shape[0],
                                                                   round(100 * (good_first + good_sec + good_third)
                                                                         / sheet_df.shape[0], 2))
              )

        print("A total of {} Good categories found in fourth choice, while first either Bad or Neutral.\n"
              "If we were to replace these, the previous {} Good choices would total {} out of {}.\n"
              "That would represent {}% of Good choices.\n".format(good_fourth,
                                                                   good_first + good_sec + good_third,
                                                                   good_first + good_sec + good_third + good_fourth,
                                                                   sheet_df.shape[0],
                                                                   round(100 * (good_first + good_sec + good_third + good_fourth)
                                                                         / sheet_df.shape[0], 2))
              )

        print("A total of {} Good categories found in fifth choice, while first either Bad or Neutral.\n"
              "If we were to replace these, the previous {} Good choices would total {} out of {}.\n"
              "That would represent {}% of Good choices.\n".format(good_fifth,
                                                                   good_first + good_sec + good_third + good_fourth,
                                                                   good_first + good_sec + good_third + good_fourth + good_fifth,
                                                                   sheet_df.shape[0],
                                                                   perc)
              )

        print("There are a total of {} topics with no Good matches ({}%)\n".format(good_neither, round(100 * good_neither / sheet_df.shape[0], 2)))


print_out()


def find_best_matches():
    """
    Given all files, find the one with most possible Good choices. Only take consideration of first Good choice found.
    :return: Writes dataframe with topic correspondence to a CSV file. Contains NA values to be resolved later with
    rules.
    """
    best_file = files[percs.index(max(percs))]

    cap = pd.read_csv("Data/Coding_Schemes/CAP.csv")

    sheet = px.load_workbook(best_file).active

    CODES = ['B', 'E', 'H', 'K', 'N']

    col_dict = {}

    # Choice values
    for col in COLS:
        col_dict[col] = [STYLE_VALUE[sheet[col + str(cell_num)].style] for cell_num in range(2, sheet.max_row + 1)]

    # Code corresponding to choices. Can include everything in the same dataframe. No need for extra data structure.
    for col in CODES:
        col_dict[col] = [sheet[col + str(cell_num)].internal_value for cell_num in range(2, sheet.max_row + 1)]

    df = pd.concat([cap.Code, pd.DataFrame(col_dict)], axis=1)

    # List containing 5 different dataframes. One for each possible outcome of first Good choice occurrence.
    first_goods = [df[df.C == 1],
                   df[(df.C != 1) & (df.F == 1)],
                   df[(df.C != 1) & (df.F != 1) & (df.I == 1)],
                   df[(df.C != 1) & (df.F != 1) & (df.I != 1) & (df.L == 1)],
                   df[(df.C != 1) & (df.F != 1) & (df.I != 1) & (df.L != 1) & (df.O == 1)]
                   ]

    code_dict = {}

    # For each dataframe above, get the corresponding Good choice and match it with its respective MAN code
    for idx in range(len(first_goods)):
        sub_df = first_goods[idx]
        code_dict.update(dict(zip(sub_df.Code, sub_df[CODES[idx]])))

    # Include the CAP code in dataframe. Left join ensures that CAP codes with no MAN code is marked as NA.
    return cap.merge(pd.DataFrame(code_dict.items(), columns=['CAP', 'MAN']), left_on='Code', right_on='CAP', how='left')[['Code', 'MAN']]


# find_best_matches().to_csv(DATA_DIR + 'cap_to_man.csv', index=False)
