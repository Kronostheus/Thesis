import openpyxl as px
import glob

WORKBOOK_DIR = "Data/Workbooks/"

COLS = ['C', 'F']

STYLE_VALUES = {"Good": 1, "Neutral": 0.5, "Bad": 0, "Normal": -1}
VALUES_STYLE = {1: "Good", 0.5: "Neutral", 0: "Bad", -1: "Normal"}

files = [file for file in glob.glob(WORKBOOK_DIR + '*.xlsx')]

to_process = [file for file in files if px.load_workbook(file).active['O2'].style == "Normal"]


def decide_style(topic_dict, cell):
    """
    An optimistic method that assumes that the best value seen for a cell should be considered. If one file had a choice
    as Neutral, but another had it as Good, the choice is considered Good for the files still to be processed.
    :param topic_dict: Dictionary<row_number, Dictionary<Topic, Choice>> Mapping of known choices for every row
    :param cell: Cell being processed at the moment
    :return: Style (choice) to be given to the cell
    """
    best_style = VALUES_STYLE[max(STYLE_VALUES[cell.style], STYLE_VALUES[topic_dict[cell.internal_value]])] \
        if cell.internal_value in topic_dict.keys() else cell.style
    return best_style


def get_cols(row_num):
    """
    Generate all cell positions (ex: C2) for the columns containing choice information
    :param row_num: Excel row number
    :return: List containing all cells with choice information from a given row
    """
    return [col + str(row_num) for col in ['C', 'F', 'I', 'L', 'O']]


last_wb_dict = {}

for file in files:

    last_wb = px.load_workbook(file).active

    for row in range(2, last_wb.max_row + 1):

        topics = [last_wb[col] for col in get_cols(row)]

        if row in last_wb_dict.keys():
            # Row already processed (not the first file being seen), get best choice for each cell
            for topic in topics:
                last_wb_dict[row][topic.internal_value] = decide_style(last_wb_dict[row], topic)
        else:
            # First time seeing this row, include all values
            last_wb_dict[row] = {topic.internal_value: topic.style for topic in topics}

for file in to_process:

    wb = px.load_workbook(file)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):

        prev_values = last_wb_dict[row]

        topics = get_cols(row)

        for topic in topics:
            # Change choice value if cell has a topic which has been previously seen
            sheet[topic].style = prev_values[sheet[topic].internal_value] \
                if sheet[topic].internal_value in prev_values.keys() else sheet[topic].style

    wb.save(file)
